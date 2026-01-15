"""Excel Export Optimization - Optimization Sheet Creation.

Refactored from excel_export.py.

Contains:
- add_optimization_sheet: Create optimization results sheet with all trials
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .excel_export import StrategySimulatorExport


class ExcelExportOptimization:
    """Helper for optimization sheet creation."""

    def __init__(self, parent: StrategySimulatorExport):
        self.parent = parent

    def add_optimization_sheet(self) -> None:
        """Add optimization results sheet."""
        if not self.parent._optimization_run:
            return

        ws = self.parent.workbook.create_sheet("Optimization")
        opt = self.parent._optimization_run

        # Header info
        ws.cell(row=1, column=1, value="Optimization Results")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Strategy: {opt.strategy_name}")
        ws.cell(row=3, column=1, value=f"Type: {opt.optimization_type}")
        ws.cell(row=4, column=1, value=f"Objective: {opt.objective_metric}")
        ws.cell(row=5, column=1, value=f"Total Trials: {opt.total_trials}")
        ws.cell(row=6, column=1, value=f"Elapsed: {opt.elapsed_seconds:.1f}s")
        ws.cell(row=7, column=1, value=f"Best Score: {opt.best_score:.4f}")

        # Best params
        ws.cell(row=8, column=1, value="Best Parameters:")
        for i, (k, v) in enumerate(opt.best_params.items()):
            ws.cell(row=8, column=2 + i, value=f"{k}={v}")

        # Trials table
        start_row = 10

        # Get all param names from first trial
        if opt.all_trials:
            param_names = list(opt.all_trials[0].parameters.keys())
            metric_names = list(opt.all_trials[0].metrics.keys())

            headers = ["Trial", "Score"] + param_names + metric_names

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = self.parent._header_font
                cell.fill = self.parent._header_fill
                cell.border = self.parent._border

            # Write trials
            for row_idx, trial in enumerate(opt.all_trials, start_row + 1):
                ws.cell(row=row_idx, column=1, value=trial.trial_number)
                ws.cell(row=row_idx, column=2, value=f"{trial.score:.4f}")

                col = 3
                for param_name in param_names:
                    value = trial.parameters.get(param_name, "")
                    ws.cell(row=row_idx, column=col, value=value)
                    col += 1

                for metric_name in metric_names:
                    value = trial.metrics.get(metric_name, "")
                    if isinstance(value, float):
                        value = f"{value:.4f}"
                    ws.cell(row=row_idx, column=col, value=value)
                    col += 1

            # Adjust widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12
