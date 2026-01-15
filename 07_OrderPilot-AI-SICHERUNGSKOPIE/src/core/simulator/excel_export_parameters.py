"""Excel Export Parameters - Parameters Sheet Creation.

Refactored from excel_export.py.

Contains:
- add_parameters_sheet: Create sheet documenting parameter definitions
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .excel_export import StrategySimulatorExport


class ExcelExportParameters:
    """Helper for parameters sheet creation."""

    def __init__(self, parent: StrategySimulatorExport):
        self.parent = parent

    def add_parameters_sheet(self) -> None:
        """Add sheet documenting parameter definitions."""
        from .strategy_params import STRATEGY_PARAMETER_REGISTRY

        ws = self.parent.workbook.create_sheet("Parameters")

        headers = [
            "Strategy",
            "Parameter",
            "Display Name",
            "Type",
            "Default",
            "Min",
            "Max",
            "Step",
            "Description",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.parent._header_font
            cell.fill = self.parent._header_fill
            cell.border = self.parent._border

        row = 2
        for strategy_name, config in STRATEGY_PARAMETER_REGISTRY.items():
            for param in config.parameters:
                data = [
                    strategy_name.value,
                    param.name,
                    param.display_name,
                    param.param_type,
                    param.default,
                    param.min_value,
                    param.max_value,
                    param.step,
                    param.description,
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.parent._border
                row += 1

        # Adjust widths
        widths = [16, 18, 16, 8, 10, 8, 8, 8, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
