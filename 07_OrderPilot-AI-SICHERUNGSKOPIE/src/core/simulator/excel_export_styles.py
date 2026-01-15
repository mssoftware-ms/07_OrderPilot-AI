"""Excel Export Styles - Style Definitions and Initialization.

Refactored from excel_export.py.

Contains:
- Openpyxl dependency check
- Style definitions (fonts, fills, borders)
"""

from __future__ import annotations

import logging

logger = logging.getLogger(__name__)

# Import openpyxl components (optional dependency)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None


def check_openpyxl_available():
    """Check if openpyxl is available.

    Raises:
        ImportError: If openpyxl is not installed
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )


def create_header_font() -> Font:
    """Create header font style."""
    return Font(bold=True, color="FFFFFF")


def create_header_fill() -> PatternFill:
    """Create header fill style."""
    return PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )


def create_positive_fill() -> PatternFill:
    """Create positive value fill style (green)."""
    return PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )


def create_negative_fill() -> PatternFill:
    """Create negative value fill style (red)."""
    return PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )


def create_border() -> Border:
    """Create cell border style."""
    return Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def create_workbook():
    """Create new openpyxl workbook.

    Returns:
        Workbook instance

    Raises:
        ImportError: If openpyxl is not installed
    """
    check_openpyxl_available()
    return openpyxl.Workbook()
