"""Excel Export Summary - Summary Sheet Creation.

Refactored from excel_export.py.

Contains:
- add_summary_sheet: Create summary sheet with all results sorted by score
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.styles import Alignment

if TYPE_CHECKING:
    from .excel_export import StrategySimulatorExport


class ExcelExportSummary:
    """Helper for summary sheet creation."""

    def __init__(self, parent: StrategySimulatorExport):
        self.parent = parent

    def add_summary_sheet(self) -> None:
        """Add summary sheet with all results, sorted by Score."""
        # Create new sheet if UI table already used active, otherwise use active
        if self.parent._ui_table_data:
            ws = self.parent.workbook.create_sheet("Summary")
        else:
            ws = self.parent.workbook.active
            ws.title = "Summary"

        headers = [
            "Strategy",
            "Symbol",
            "Trades",
            "Win %",
            "PF",
            "P&L €",
            "P&L %",
            "DD %",
            "Score",
            "Parameters",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.parent._header_font
            cell.fill = self.parent._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.parent._border

        # Sort results by Score (descending)
        sorted_results = sorted(
            self.parent._results,
            key=lambda r: self.parent._calculate_score(r),
            reverse=True
        )

        # Write data
        for row_idx, result in enumerate(sorted_results, 2):
            # Show ALL parameters
            params_str = ", ".join(
                f"{k}={v}" for k, v in result.parameters.items()
            )

            score = self.parent._calculate_score(result)

            data = [
                result.strategy_name,
                result.symbol,
                result.total_trades,
                f"{result.win_rate * 100:.1f}",
                f"{result.profit_factor:.2f}",
                f"{result.total_pnl:.2f}",
                f"{result.total_pnl_pct:.2f}",
                f"{result.max_drawdown_pct:.2f}",
                score,
                params_str,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.parent._border

                # Color P&L cells
                if col == 6:  # P&L €
                    try:
                        pnl_val = float(str(value).replace(",", ""))
                        if pnl_val > 0:
                            cell.fill = self.parent._positive_fill
                        elif pnl_val < 0:
                            cell.fill = self.parent._negative_fill
                    except (ValueError, TypeError):
                        pass

                # Color Score cells
                if col == 9:  # Score
                    try:
                        score_val = int(value)
                        if score_val > 0:
                            cell.fill = self.parent._positive_fill
                        elif score_val < 0:
                            cell.fill = self.parent._negative_fill
                    except (ValueError, TypeError):
                        pass

        # Adjust column widths
        ws.column_dimensions["A"].width = 16  # Strategy
        ws.column_dimensions["B"].width = 12  # Symbol
        ws.column_dimensions["C"].width = 8   # Trades
        ws.column_dimensions["D"].width = 8   # Win %
        ws.column_dimensions["E"].width = 8   # PF
        ws.column_dimensions["F"].width = 12  # P&L €
        ws.column_dimensions["G"].width = 10  # P&L %
        ws.column_dimensions["H"].width = 8   # DD %
        ws.column_dimensions["I"].width = 10  # Score
        ws.column_dimensions["J"].width = 100 # Parameters (wide for all params)
