"""Excel Export Utilities for Strategy Simulation.

Exports simulation results to Excel (.xlsx) files with multiple sheets.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from .result_types import SimulationResult, OptimizationRun, TradeRecord

logger = logging.getLogger(__name__)

# Import openpyxl components (optional dependency)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None


class StrategySimulatorExport:
    """Export strategy simulation results to Excel.

    Creates a multi-sheet Excel file with:
    - UI Table: Exact copy of UI results table (first sheet)
    - Summary: Overview of all simulation results
    - Optimization: All optimization trials (if applicable)
    - Trades_{Strategy}: Detailed trades for each result
    - Parameters: Parameter definitions and ranges
    """

    def __init__(self, filepath: Path | str):
        """Initialize exporter.

        Args:
            filepath: Path to save the Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        self.filepath = Path(filepath)
        self.workbook = openpyxl.Workbook()
        self._results: list[SimulationResult] = []
        self._optimization_run: OptimizationRun | None = None
        self._ui_table_data: list[list[str]] | None = None

        # Styles
        self._header_font = Font(bold=True, color="FFFFFF")
        self._header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self._positive_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        self._negative_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        self._border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def add_results(self, results: list[SimulationResult]) -> None:
        """Add simulation results to export.

        Args:
            results: List of SimulationResult objects
        """
        self._results.extend(results)

    def add_optimization_run(self, opt_run: OptimizationRun) -> None:
        """Add optimization run to export.

        Args:
            opt_run: OptimizationRun with all trials
        """
        self._optimization_run = opt_run

    def add_ui_table_data(self, table_data: list[list[str]]) -> None:
        """Add UI table data for export as first sheet.

        Args:
            table_data: List of rows, each row is a list of cell values.
                        First row should be headers.
        """
        self._ui_table_data = table_data

    def _calculate_score(self, result) -> int:
        """Calculate performance score from -1000 to +1000."""
        if getattr(result, "entry_only", False):
            entry_score = result.entry_score or 0.0
            return int(round(entry_score))
        pnl_pct = result.total_pnl_pct
        return int(max(-1000, min(1000, pnl_pct * 10)))

    def _format_entry_points(self, points: list[tuple] | None) -> str:
        if not points:
            return ""
        formatted = []
        for item in points:
            if len(item) == 3:
                price, ts, _score = item
            else:
                price, ts = item
            formatted.append(f"{price:.3f}/{ts.strftime('%H:%M:%S')}")
        return ";".join(formatted)

    def _format_params_with_entry(self, result: SimulationResult, params_str: str) -> str:
        if not getattr(result, "entry_only", False):
            return params_str
        ls_value = getattr(result, "entry_side", None) or "long"
        entry_display = self._format_entry_points(getattr(result, "entry_points", None))
        if not entry_display:
            entry_price = getattr(result, "entry_best_price", None)
            entry_time = getattr(result, "entry_best_time", None)
            if entry_price is not None and entry_time is not None:
                entry_display = f"{entry_price:.3f}/{entry_time.strftime('%H:%M:%S')}"
        parts = [f"LS={ls_value}"]
        if entry_display:
            parts.append(f"EP={entry_display}")
        prefix = ", ".join(parts)
        return f"{prefix}, {params_str}" if params_str else prefix

    def add_ui_table_sheet(self) -> None:
        """Add UI table as first sheet (exact copy of displayed table)."""
        if not self._ui_table_data:
            return

        ws = self.workbook.active
        ws.title = "Ergebnisse"

        headers = [
            "Strategy",
            "Trades",
            "Win %",
            "PF",
            "P&L €",
            "P&L %",
            "DD %",
            "Score",
            "Objective",
            "Parameters",
        ]

        objective_by_key: dict[tuple[str, str], str] = {}
        if self._ui_table_data:
            for row in self._ui_table_data:
                if len(row) >= 10:
                    strategy = row[0]
                    params = row[9]
                    objective = row[8]
                    if strategy and params and objective:
                        objective_by_key[(strategy, params)] = objective

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._border

        # Write data rows
        for row_idx, row_data in enumerate(self._ui_table_data, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self._border

                # Color P&L € cells (column 5)
                if col == 5:
                    try:
                        pnl_val = float(str(value).replace(",", ""))
                        if pnl_val > 0:
                            cell.fill = self._positive_fill
                        elif pnl_val < 0:
                            cell.fill = self._negative_fill
                    except (ValueError, TypeError):
                        pass

                # Color Score cells (column 8)
                if col == 8:
                    try:
                        score_val = int(float(value))
                        if score_val > 0:
                            cell.fill = self._positive_fill
                        elif score_val < 0:
                            cell.fill = self._negative_fill
                    except (ValueError, TypeError):
                        pass

        # Adjust column widths
        ws.column_dimensions["A"].width = 16  # Strategy
        ws.column_dimensions["B"].width = 8   # Trades
        ws.column_dimensions["C"].width = 8   # Win %
        ws.column_dimensions["D"].width = 8   # PF
        ws.column_dimensions["E"].width = 12  # P&L €
        ws.column_dimensions["F"].width = 10  # P&L %
        ws.column_dimensions["G"].width = 8   # DD %
        ws.column_dimensions["H"].width = 10  # Score
        ws.column_dimensions["I"].width = 12  # Objective
        ws.column_dimensions["J"].width = 100 # Parameters

        # Add parameter legend below data
        self._add_parameter_legend(ws, len(self._ui_table_data) + 4)

    def _add_parameter_legend(self, ws, start_row: int) -> None:
        """Add parameter legend below the results table.

        Args:
            ws: Worksheet to add legend to
            start_row: Row to start the legend
        """
        from .strategy_params import STRATEGY_PARAMETER_REGISTRY, StrategyName

        # Legend header
        legend_header_font = Font(bold=True, size=12)
        legend_subheader_font = Font(bold=True, size=10, color="2F5496")

        cell = ws.cell(row=start_row, column=1, value="PARAMETER-LEGENDE")
        cell.font = legend_header_font

        current_row = start_row + 2

        # Add legend for each strategy
        for strategy_name in StrategyName:
            config = STRATEGY_PARAMETER_REGISTRY.get(strategy_name)
            if not config:
                continue

            # Strategy header
            cell = ws.cell(
                row=current_row, column=1,
                value=f"● {config.display_name} ({strategy_name.value})"
            )
            cell.font = legend_subheader_font
            current_row += 1

            # Parameter table headers
            param_headers = ["Parameter", "Beschreibung", "Typ", "Default", "Min", "Max"]
            for col, header in enumerate(param_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(
                    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
                )
            current_row += 1

            # Parameter rows
            for param in config.parameters:
                ws.cell(row=current_row, column=1, value=param.name)
                ws.cell(row=current_row, column=2, value=param.description)
                ws.cell(row=current_row, column=3, value=param.param_type)
                ws.cell(row=current_row, column=4, value=str(param.default))
                ws.cell(row=current_row, column=5, value=str(param.min_value) if param.min_value is not None else "-")
                ws.cell(row=current_row, column=6, value=str(param.max_value) if param.max_value is not None else "-")
                current_row += 1

            # Empty row between strategies
            current_row += 1

        # Adjust column widths for legend
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 50)

    def add_summary_sheet(self) -> None:
        """Add summary sheet with all results, sorted by Score."""
        # Create new sheet if UI table already used active, otherwise use active
        if self._ui_table_data:
            ws = self.workbook.create_sheet("Summary")
        else:
            ws = self.workbook.active
            ws.title = "Summary"

        objective_by_key: dict[tuple[str, str], str] = {}
        if self._ui_table_data:
            for row in self._ui_table_data:
                if len(row) >= 10:
                    strategy = row[0]
                    params = row[9]
                    objective = row[8]
                    if strategy and params and objective:
                        objective_by_key[(strategy, params)] = objective

        headers = [
            "Strategy",
            "Symbol",
            "Trades",
            "Win %",
            "PF",
            "P&L €",
            "P&L %",
            "DD %",
            "Score",
            "Objective",
            "Parameters",
        ]

        objective_labels = {
            "score": "Score",
            "entry_score": "Entry Score",
            "total_pnl_pct": "P&L %",
            "profit_factor": "PF",
            "sharpe_ratio": "Sharpe",
            "win_rate": "Win %",
            "max_drawdown_pct": "Max DD",
        }

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._border

        # Sort results by Score (descending)
        sorted_results = sorted(
            self._results,
            key=lambda r: self._calculate_score(r),
            reverse=True
        )

        # Write data
        for row_idx, result in enumerate(sorted_results, 2):
            # Show ALL parameters
            params_str = ", ".join(
                f"{k}={v}" for k, v in result.parameters.items()
            )
            params_str = self._format_params_with_entry(result, params_str)

            objective_label = objective_by_key.get((result.strategy_name, params_str))
            if not objective_label:
                if self._optimization_run:
                    metric = self._optimization_run.objective_metric
                    objective_label = objective_labels.get(metric, metric)
                else:
                    objective_label = "Manual"

            score = self._calculate_score(result)

            data = [
                result.strategy_name,
                result.symbol,
                result.total_trades,
                f"{result.win_rate * 100:.1f}",
                f"{result.profit_factor:.2f}",
                f"{result.total_pnl:.2f}",
                f"{result.total_pnl_pct:.2f}",
                f"{result.max_drawdown_pct:.2f}",
                score,
                objective_label,
                params_str,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self._border

                # Color P&L cells
                if col == 6:  # P&L €
                    try:
                        pnl_val = float(str(value).replace(",", ""))
                        if pnl_val > 0:
                            cell.fill = self._positive_fill
                        elif pnl_val < 0:
                            cell.fill = self._negative_fill
                    except (ValueError, TypeError):
                        pass

                # Color Score cells
                if col == 9:  # Score
                    try:
                        score_val = int(value)
                        if score_val > 0:
                            cell.fill = self._positive_fill
                        elif score_val < 0:
                            cell.fill = self._negative_fill
                    except (ValueError, TypeError):
                        pass

        # Adjust column widths
        ws.column_dimensions["A"].width = 16  # Strategy
        ws.column_dimensions["B"].width = 12  # Symbol
        ws.column_dimensions["C"].width = 8   # Trades
        ws.column_dimensions["D"].width = 8   # Win %
        ws.column_dimensions["E"].width = 8   # PF
        ws.column_dimensions["F"].width = 12  # P&L €
        ws.column_dimensions["G"].width = 10  # P&L %
        ws.column_dimensions["H"].width = 8   # DD %
        ws.column_dimensions["I"].width = 10  # Score
        ws.column_dimensions["J"].width = 12  # Objective
        ws.column_dimensions["K"].width = 100 # Parameters (wide for all params)

    def add_optimization_sheet(self) -> None:
        """Add optimization results sheet."""
        if not self._optimization_run:
            return

        ws = self.workbook.create_sheet("Optimization")
        opt = self._optimization_run

        # Header info
        ws.cell(row=1, column=1, value="Optimization Results")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Strategy: {opt.strategy_name}")
        ws.cell(row=3, column=1, value=f"Type: {opt.optimization_type}")
        ws.cell(row=4, column=1, value=f"Objective: {opt.objective_metric}")
        ws.cell(row=5, column=1, value=f"Total Trials: {opt.total_trials}")
        ws.cell(row=6, column=1, value=f"Elapsed: {opt.elapsed_seconds:.1f}s")
        ws.cell(row=7, column=1, value=f"Best Score: {opt.best_score:.4f}")

        # Best params
        ws.cell(row=8, column=1, value="Best Parameters:")
        for i, (k, v) in enumerate(opt.best_params.items()):
            ws.cell(row=8, column=2 + i, value=f"{k}={v}")

        # Trials table
        start_row = 10

        # Get all param names from first trial
        if opt.all_trials:
            param_names = list(opt.all_trials[0].parameters.keys())
            metric_names = list(opt.all_trials[0].metrics.keys())

            headers = ["Trial", "Score"] + param_names + metric_names

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.border = self._border

            # Write trials
            for row_idx, trial in enumerate(opt.all_trials, start_row + 1):
                ws.cell(row=row_idx, column=1, value=trial.trial_number)
                ws.cell(row=row_idx, column=2, value=f"{trial.score:.4f}")

                col = 3
                for param_name in param_names:
                    value = trial.parameters.get(param_name, "")
                    ws.cell(row=row_idx, column=col, value=value)
                    col += 1

                for metric_name in metric_names:
                    value = trial.metrics.get(metric_name, "")
                    if isinstance(value, float):
                        value = f"{value:.4f}"
                    ws.cell(row=row_idx, column=col, value=value)
                    col += 1

            # Adjust widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

    def add_trades_sheet(self, result: SimulationResult, sheet_name: str = None) -> None:
        """Add detailed trades sheet for a specific result.

        Args:
            result: SimulationResult with trades
            sheet_name: Optional custom sheet name
        """
        name = sheet_name or f"Trades_{result.strategy_name[:20]}"
        # Ensure unique sheet name
        existing = [ws.title for ws in self.workbook.worksheets]
        if name in existing:
            name = f"{name}_{len(existing)}"

        ws = self.workbook.create_sheet(name)

        headers = [
            "Entry Time",
            "Exit Time",
            "Side",
            "Size",
            "Entry Price",
            "Exit Price",
            "Stop Loss",
            "Take Profit",
            "P&L",
            "P&L %",
            "Exit Reason",
            "Duration (min)",
            "Commission",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border

        # Write trades
        for row_idx, trade in enumerate(result.trades, 2):
            duration_min = trade.duration_seconds / 60 if trade.duration_seconds else 0

            data = [
                trade.entry_time.strftime("%Y-%m-%d %H:%M"),
                trade.exit_time.strftime("%Y-%m-%d %H:%M"),
                trade.side,
                f"{trade.size:.4f}",
                f"{trade.entry_price:.4f}",
                f"{trade.exit_price:.4f}",
                f"{trade.stop_loss:.4f}" if trade.stop_loss else "N/A",
                f"{trade.take_profit:.4f}" if trade.take_profit else "N/A",
                f"{trade.pnl:.2f}",
                f"{trade.pnl_pct * 100:.2f}",
                trade.exit_reason,
                f"{duration_min:.1f}",
                f"{trade.commission:.4f}",
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self._border

                # Color P&L cells
                if col == 9:  # P&L column
                    try:
                        pnl_val = float(str(value).replace(",", ""))
                        if pnl_val > 0:
                            cell.fill = self._positive_fill
                        elif pnl_val < 0:
                            cell.fill = self._negative_fill
                    except ValueError:
                        pass

        # Adjust column widths
        widths = [18, 18, 8, 10, 12, 12, 12, 12, 10, 10, 14, 12, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def add_parameters_sheet(self) -> None:
        """Add sheet documenting parameter definitions."""
        from .strategy_params import STRATEGY_PARAMETER_REGISTRY

        ws = self.workbook.create_sheet("Parameters")

        headers = [
            "Strategy",
            "Parameter",
            "Display Name",
            "Type",
            "Default",
            "Min",
            "Max",
            "Step",
            "Description",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border

        row = 2
        for strategy_name, config in STRATEGY_PARAMETER_REGISTRY.items():
            for param in config.parameters:
                data = [
                    strategy_name.value,
                    param.name,
                    param.display_name,
                    param.param_type,
                    param.default,
                    param.min_value,
                    param.max_value,
                    param.step,
                    param.description,
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self._border
                row += 1

        # Adjust widths
        widths = [16, 18, 16, 8, 10, 8, 8, 8, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def save(self) -> Path:
        """Save workbook and return path.

        Returns:
            Path to saved file
        """
        # Build sheets - UI table first if available
        if self._ui_table_data:
            self.add_ui_table_sheet()

        self.add_summary_sheet()

        if self._optimization_run:
            self.add_optimization_sheet()

        # Add trades sheets for top results (max 5)
        for result in self._results[:5]:
            if result.trades:
                self.add_trades_sheet(result)

        self.add_parameters_sheet()

        # Save
        self.workbook.save(self.filepath)
        logger.info(f"Saved simulation results to {self.filepath}")

        return self.filepath


def export_simulation_results(
    results: list[SimulationResult],
    filepath: Path | str,
    optimization_run: OptimizationRun | None = None,
    ui_table_data: list[list[str]] | None = None,
) -> Path:
    """Convenience function to export simulation results.

    Args:
        results: List of simulation results
        filepath: Path to save Excel file
        optimization_run: Optional optimization run with all trials
        ui_table_data: Optional list of table rows from UI (exported as first sheet)

    Returns:
        Path to saved file
    """
    exporter = StrategySimulatorExport(filepath)
    exporter.add_results(results)

    if optimization_run:
        exporter.add_optimization_run(optimization_run)

    if ui_table_data:
        exporter.add_ui_table_data(ui_table_data)

    return exporter.save()
