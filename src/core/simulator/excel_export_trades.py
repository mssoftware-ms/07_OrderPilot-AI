"""Excel Export Trades - Trades Sheet Creation.

Refactored from excel_export.py.

Contains:
- add_trades_sheet: Create detailed trades sheet for a specific result
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .excel_export import StrategySimulatorExport
    from .result_types import SimulationResult


class ExcelExportTrades:
    """Helper for trades sheet creation."""

    def __init__(self, parent: StrategySimulatorExport):
        self.parent = parent

    def add_trades_sheet(self, result: SimulationResult, sheet_name: str = None) -> None:
        """Add detailed trades sheet for a specific result.

        Args:
            result: SimulationResult with trades
            sheet_name: Optional custom sheet name
        """
        name = sheet_name or f"Trades_{result.strategy_name[:20]}"
        # Ensure unique sheet name
        existing = [ws.title for ws in self.parent.workbook.worksheets]
        if name in existing:
            name = f"{name}_{len(existing)}"

        ws = self.parent.workbook.create_sheet(name)

        headers = [
            "Entry Time",
            "Exit Time",
            "Side",
            "Size",
            "Entry Price",
            "Exit Price",
            "Stop Loss",
            "Take Profit",
            "P&L",
            "P&L %",
            "Exit Reason",
            "Duration (min)",
            "Commission",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.parent._header_font
            cell.fill = self.parent._header_fill
            cell.border = self.parent._border

        # Write trades
        for row_idx, trade in enumerate(result.trades, 2):
            duration_min = trade.duration_seconds / 60 if trade.duration_seconds else 0

            data = [
                trade.entry_time.strftime("%Y-%m-%d %H:%M"),
                trade.exit_time.strftime("%Y-%m-%d %H:%M"),
                trade.side,
                f"{trade.size:.4f}",
                f"{trade.entry_price:.4f}",
                f"{trade.exit_price:.4f}",
                f"{trade.stop_loss:.4f}" if trade.stop_loss else "N/A",
                f"{trade.take_profit:.4f}" if trade.take_profit else "N/A",
                f"{trade.pnl:.2f}",
                f"{trade.pnl_pct * 100:.2f}",
                trade.exit_reason,
                f"{duration_min:.1f}",
                f"{trade.commission:.4f}",
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.parent._border

                # Color P&L cells
                if col == 9:  # P&L column
                    try:
                        pnl_val = float(str(value).replace(",", ""))
                        if pnl_val > 0:
                            cell.fill = self.parent._positive_fill
                        elif pnl_val < 0:
                            cell.fill = self.parent._negative_fill
                    except ValueError:
                        pass

        # Adjust column widths
        widths = [18, 18, 8, 10, 12, 12, 12, 12, 10, 10, 14, 12, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
