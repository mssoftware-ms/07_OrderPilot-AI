"""Excel Export UI Table - UI Table Sheet Creation.

Refactored from excel_export.py.

Contains:
- add_ui_table_sheet: Create UI results table sheet
- _add_parameter_legend: Add parameter legend below table
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.styles import Font, PatternFill, Alignment

if TYPE_CHECKING:
    from .excel_export import StrategySimulatorExport


class ExcelExportUITable:
    """Helper for UI table sheet creation."""

    def __init__(self, parent: StrategySimulatorExport):
        self.parent = parent

    def add_ui_table_sheet(self) -> None:
        """Add UI table as first sheet (exact copy of displayed table)."""
        if not self.parent._ui_table_data:
            return

        ws = self.parent.workbook.active
        ws.title = "Ergebnisse"

        headers = [
            "Strategy",
            "Trades",
            "Win %",
            "PF",
            "P&L €",
            "P&L %",
            "DD %",
            "Score",
            "Parameters",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.parent._header_font
            cell.fill = self.parent._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.parent._border

        # Write data rows
        for row_idx, row_data in enumerate(self.parent._ui_table_data, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.parent._border

                # Color P&L € cells (column 5)
                if col == 5:
                    try:
                        pnl_val = float(str(value).replace(",", ""))
                        if pnl_val > 0:
                            cell.fill = self.parent._positive_fill
                        elif pnl_val < 0:
                            cell.fill = self.parent._negative_fill
                    except (ValueError, TypeError):
                        pass

                # Color Score cells (column 8)
                if col == 8:
                    try:
                        score_val = int(float(value))
                        if score_val > 0:
                            cell.fill = self.parent._positive_fill
                        elif score_val < 0:
                            cell.fill = self.parent._negative_fill
                    except (ValueError, TypeError):
                        pass

        # Adjust column widths
        ws.column_dimensions["A"].width = 16  # Strategy
        ws.column_dimensions["B"].width = 8   # Trades
        ws.column_dimensions["C"].width = 8   # Win %
        ws.column_dimensions["D"].width = 8   # PF
        ws.column_dimensions["E"].width = 12  # P&L €
        ws.column_dimensions["F"].width = 10  # P&L %
        ws.column_dimensions["G"].width = 8   # DD %
        ws.column_dimensions["H"].width = 10  # Score
        ws.column_dimensions["I"].width = 100 # Parameters

        # Add parameter legend below data
        self.add_parameter_legend(ws, len(self.parent._ui_table_data) + 4)

    def add_parameter_legend(self, ws, start_row: int) -> None:
        """Add parameter legend below the results table.

        Args:
            ws: Worksheet to add legend to
            start_row: Row to start the legend
        """
        from .strategy_params import STRATEGY_PARAMETER_REGISTRY, StrategyName

        # Legend header
        legend_header_font = Font(bold=True, size=12)
        legend_subheader_font = Font(bold=True, size=10, color="2F5496")

        cell = ws.cell(row=start_row, column=1, value="PARAMETER-LEGENDE")
        cell.font = legend_header_font

        current_row = start_row + 2

        # Add legend for each strategy
        for strategy_name in StrategyName:
            config = STRATEGY_PARAMETER_REGISTRY.get(strategy_name)
            if not config:
                continue

            # Strategy header
            cell = ws.cell(
                row=current_row, column=1,
                value=f"● {config.display_name} ({strategy_name.value})"
            )
            cell.font = legend_subheader_font
            current_row += 1

            # Parameter table headers
            param_headers = ["Parameter", "Beschreibung", "Typ", "Default", "Min", "Max"]
            for col, header in enumerate(param_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(
                    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
                )
            current_row += 1

            # Parameter rows
            for param in config.parameters:
                ws.cell(row=current_row, column=1, value=param.name)
                ws.cell(row=current_row, column=2, value=param.description)
                ws.cell(row=current_row, column=3, value=param.param_type)
                ws.cell(row=current_row, column=4, value=str(param.default))
                ws.cell(row=current_row, column=5, value=str(param.min_value) if param.min_value is not None else "-")
                ws.cell(row=current_row, column=6, value=str(param.max_value) if param.max_value is not None else "-")
                current_row += 1

            # Empty row between strategies
            current_row += 1

        # Adjust column widths for legend
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 50)
