from __future__ import annotations

import asyncio
import logging
from datetime import datetime
from decimal import Decimal

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QPainter, QLinearGradient
from PyQt6.QtWidgets import (
    QFormLayout, QGroupBox, QHBoxLayout, QHeaderView, QLabel,
    QProgressBar, QPushButton, QSplitter, QTableWidget, QVBoxLayout, QWidget,
    QMessageBox, QPlainTextEdit, QFileDialog
)

from .bot_sltp_progressbar import SLTPProgressBar

logger = logging.getLogger(__name__)

class BotUISignalsActionsMixin:
    """Signal and position actions"""

    def _on_clear_selected_signal(self) -> None:
        """Löscht die ausgewählte Zeile aus der Signals-Tabelle."""
        selected_rows = self.signals_table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # Von unten nach oben löschen, um Index-Verschiebung zu vermeiden
        rows_to_delete = sorted([idx.row() for idx in selected_rows], reverse=True)

        # Map table rows (latest on top) back to _signal_history indices
        history_len = len(self._signal_history)
        indices_to_delete = {
            history_len - 1 - row for row in rows_to_delete
            if 0 <= history_len - 1 - row < history_len
        }

        # Remove from history and UI
        if indices_to_delete:
            self._signal_history = [
                sig for i, sig in enumerate(self._signal_history)
                if i not in indices_to_delete
            ]
            self._save_signal_history()

        for row in rows_to_delete:
            self.signals_table.removeRow(row)

        # Refresh to ensure P&L / selection stays consistent
        self._update_signals_table()

    def _on_clear_all_signals(self) -> None:
        """Löscht alle Zeilen aus der Signals-Tabelle."""
        if self.signals_table.rowCount() == 0:
            return

        reply = QMessageBox.question(
            self,
            "Alle Signale löschen",
            f"Alle {self.signals_table.rowCount()} Signale aus der Tabelle löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            self._signal_history.clear()
            self._save_signal_history()
            self.signals_table.setRowCount(0)
            self._update_signals_table()

    def _on_sell_position_clicked(self) -> None:
        """Issue #11: Sell current position with limit order 0.05% below current price."""
        # Find active position
        active_sig = None
        if hasattr(self, '_find_active_signal'):
            active_sig = self._find_active_signal()

        if not active_sig:
            QMessageBox.warning(
                self,
                "Keine offene Position",
                "Es gibt keine offene Position zum Verkaufen.",
            )
            return

        # Get current price
        current_price = active_sig.get("current_price", 0)
        if current_price <= 0:
            current_price = active_sig.get("price", 0)

        if current_price <= 0:
            QMessageBox.warning(
                self,
                "Kein Kurs verfügbar",
                "Aktueller Kurs ist nicht verfügbar.",
            )
            return

        # Calculate limit price (0.05% below current for long, above for short)
        side = active_sig.get("side", "long").lower()
        current_price_decimal = Decimal(str(current_price))
        if side == "long":
            limit_price = current_price_decimal * Decimal("0.9995")  # 0.05% below
            order_side = "SELL"
        else:
            limit_price = current_price_decimal * Decimal("1.0005")  # 0.05% above for short
            order_side = "BUY"

        quantity = active_sig.get("quantity", 0)
        symbol = active_sig.get("symbol", "")

        if not symbol:
            # Try to get symbol from chart widget
            if hasattr(self, 'chart_widget') and hasattr(self.chart_widget, 'current_symbol'):
                symbol = self.chart_widget.current_symbol

        if quantity <= 0:
            QMessageBox.warning(
                self,
                "Keine Positionsgröße",
                "Die Positionsgröße ist nicht verfügbar.",
            )
            return

        # Confirmation dialog
        reply = QMessageBox.question(
            self,
            "Position verkaufen",
            f"Position verkaufen?\n\n"
            f"Symbol: {symbol}\n"
            f"Side: {order_side}\n"
            f"Menge: {quantity:.6f}\n"
            f"Aktueller Kurs: {current_price:.4f}\n"
            f"Limit-Preis: {limit_price:.4f} (0.05% {'unter' if side == 'long' else 'über'} Kurs)\n",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        # Execute the sell
        asyncio.create_task(self._execute_sell_order(
            symbol=symbol,
            side=order_side,
            quantity=quantity,
            limit_price=limit_price,
            signal=active_sig,
        ))

    async def _execute_sell_order(
        self,
        symbol: str,
        side: str,
        quantity: float,
        limit_price: Decimal,
        signal: dict,
    ) -> None:
        """Execute the sell order via adapter.

        Issue #11: Places a limit order to close the position.

        Args:
            symbol: Trading symbol
            side: Order side (SELL for long, BUY for short)
            quantity: Position quantity
            limit_price: Limit price for the order
            signal: The signal dict containing position info
        """
        try:
            # Get adapter
            adapter = getattr(self, '_bitunix_adapter', None)
            if not adapter:
                # Try from bitunix widget
                bitunix_widget = getattr(self, '_bitunix_widget', None)
                if bitunix_widget:
                    adapter = getattr(bitunix_widget, 'adapter', None)

            if not adapter:
                logger.error("No adapter available for selling position")
                QMessageBox.critical(
                    self,
                    "Fehler",
                    "Kein Broker-Adapter verfügbar.",
                )
                return

            # Import order types
            from src.core.broker.broker_types import OrderRequest
            from src.database.models import OrderSide, OrderType as DBOrderType

            # Create order request
            order = OrderRequest(
                symbol=symbol,
                side=OrderSide.SELL if side == "SELL" else OrderSide.BUY,
                quantity=Decimal(str(quantity)),
                order_type=DBOrderType.LIMIT,
                limit_price=Decimal(str(limit_price)),
            )

            logger.info(
                f"Issue #11: Placing limit {side} order: {symbol} qty={quantity} @ {limit_price}"
            )

            # Place order
            response = await adapter.place_order(order)

            if response and response.broker_order_id:
                logger.info(f"Issue #11: Order placed successfully: {response.broker_order_id}")

                # Update signal status
                signal["status"] = "CLOSING"
                signal["exit_order_id"] = response.broker_order_id
                if hasattr(self, '_save_signal_history'):
                    self._save_signal_history()
                if hasattr(self, '_update_signals_table'):
                    self._update_signals_table()

                # Add to KI log
                if hasattr(self, '_add_ki_log_entry'):
                    self._add_ki_log_entry(
                        "SELL",
                        f"Limit-Order platziert: {symbol} {side} {quantity:.6f} @ {limit_price:.4f}"
                    )

                # Issue #13: Remove position lines from chart when selling
                if hasattr(self, 'chart_widget'):
                    try:
                        self.chart_widget.remove_stop_line("initial_stop")
                        self.chart_widget.remove_stop_line("trailing_stop")
                        self.chart_widget.remove_stop_line("entry_line")
                        if hasattr(self, '_add_ki_log_entry'):
                            self._add_ki_log_entry("CHART", "Linien entfernt (Verkauf eingeleitet)")
                    except Exception as e:
                        logger.error(f"Error removing chart lines: {e}")

                QMessageBox.information(
                    self,
                    "Order platziert",
                    f"Limit-Order wurde platziert.\n\n"
                    f"Order ID: {response.broker_order_id}\n"
                    f"Preis: {limit_price:.4f}",
                )
            else:
                logger.error("Issue #11: Order placement failed - no order ID returned")
                QMessageBox.warning(
                    self,
                    "Order fehlgeschlagen",
                    "Die Order konnte nicht platziert werden. Siehe Logs für Details.",
                )

        except Exception as e:
            logger.exception(f"Issue #11: Failed to place sell order: {e}")
            QMessageBox.critical(
                self,
                "Fehler",
                f"Fehler beim Platzieren der Order:\n{str(e)}",
            )

    def _update_sell_button_state(self) -> None:
        """Update the sell button enabled state based on open positions.

        Issue #11: Button is enabled only when there's an open position.
        Issue #18: Also updates draw_chart_elements_btn state.
        """
        has_open_position = False
        if hasattr(self, '_find_active_signal'):
            active_sig = self._find_active_signal()
            has_open_position = active_sig is not None

        if hasattr(self, 'sell_position_btn'):
            self.sell_position_btn.setEnabled(has_open_position)

        # Issue #18: Update chart elements button state
        if hasattr(self, 'draw_chart_elements_btn'):
            self.draw_chart_elements_btn.setEnabled(has_open_position)

    def _export_signals_to_xlsx(self) -> None:
        """Export signals table to XLSX file."""
        try:
            from datetime import datetime
            from PyQt6.QtWidgets import QFileDialog, QMessageBox

            # Ask user for save location
            default_filename = f"signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Export Signals to XLSX",
                default_filename,
                "Excel Files (*.xlsx)"
            )

            if not filepath:
                return  # User cancelled

            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                QMessageBox.warning(
                    self,
                    "Module nicht gefunden",
                    "openpyxl ist nicht installiert.\n\nBitte installieren Sie es mit:\npip install openpyxl"
                )
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Signals"

            # Get headers from table
            headers = []
            for col in range(self.signals_table.columnCount()):
                if not self.signals_table.isColumnHidden(col):
                    header_item = self.signals_table.horizontalHeaderItem(col)
                    headers.append(header_item.text() if header_item else f"Column {col}")

            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Write data
            visible_col_mapping = []
            for col in range(self.signals_table.columnCount()):
                if not self.signals_table.isColumnHidden(col):
                    visible_col_mapping.append(col)

            for row_idx in range(self.signals_table.rowCount()):
                for excel_col_idx, table_col_idx in enumerate(visible_col_mapping, start=1):
                    item = self.signals_table.item(row_idx, table_col_idx)
                    if item:
                        ws.cell(row=row_idx + 2, column=excel_col_idx, value=item.text())

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save workbook
            wb.save(filepath)

            QMessageBox.information(
                self,
                "Export erfolgreich",
                f"Signals wurden erfolgreich exportiert:\n{filepath}"
            )

        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(
                self,
                "Export Fehler",
                f"Fehler beim Exportieren der Signale:\n{str(e)}"
            )

