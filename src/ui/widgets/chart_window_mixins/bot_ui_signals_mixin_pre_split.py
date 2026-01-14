from __future__ import annotations

import asyncio
import logging
from datetime import datetime
from decimal import Decimal

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QPainter, QLinearGradient
from PyQt6.QtWidgets import (
    QFormLayout, QGroupBox, QHBoxLayout, QHeaderView, QLabel,
    QProgressBar, QPushButton, QSplitter, QTableWidget, QVBoxLayout, QWidget,
    QMessageBox, QPlainTextEdit, QFileDialog
)

logger = logging.getLogger(__name__)


class SLTPProgressBar(QProgressBar):
    """Custom progress bar showing price position between SL and TP."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setRange(0, 100)
        self.setValue(50)
        self.setTextVisible(True)
        self.setMinimumHeight(28)
        self.setMaximumHeight(32)
        self._sl_price = 0.0
        self._tp_price = 0.0
        self._current_price = 0.0
        self._entry_price = 0.0
        self._side = "long"

    def set_prices(self, entry: float, sl: float, tp: float, current: float, side: str = "long"):
        """Update the bar with new prices."""
        self._entry_price = entry
        self._sl_price = sl
        self._tp_price = tp
        self._current_price = current
        self._side = side.lower()

        if sl <= 0 or tp <= 0 or sl == tp:
            self.setValue(50)
            self.setFormat("No SL/TP set")
            return

        # Calculate position as percentage (SL=0%, TP=100%)
        total_range = tp - sl
        if total_range != 0:
            position = ((current - sl) / total_range) * 100
            position = max(0, min(100, position))  # Clamp to 0-100
        else:
            position = 50

        self.setValue(int(position))

        # Format text
        pnl_pct = 0
        if entry > 0:
            if self._side == "long":
                pnl_pct = ((current - entry) / entry) * 100
            else:
                pnl_pct = ((entry - current) / entry) * 100

        sign = "+" if pnl_pct >= 0 else ""
        self.setFormat(f"SL: {sl:.2f} | Current: {current:.2f} ({sign}{pnl_pct:.2f}%) | TP: {tp:.2f}")
        self.update()

    def paintEvent(self, event):
        """Custom paint for gradient background."""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        rect = self.rect()
        rect.adjust(1, 1, -1, -1)

        # Background gradient: red (SL) -> orange (entry) -> green (TP)
        gradient = QLinearGradient(rect.left(), 0, rect.right(), 0)
        gradient.setColorAt(0.0, QColor("#ef5350"))   # Red (SL)
        gradient.setColorAt(0.5, QColor("#ff9800"))   # Orange (Entry)
        gradient.setColorAt(1.0, QColor("#26a69a"))   # Green (TP)

        painter.setBrush(gradient)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(rect, 4, 4)

        # Draw current position marker
        if self._sl_price > 0 and self._tp_price > 0 and self._tp_price != self._sl_price:
            total_range = self._tp_price - self._sl_price
            pos_ratio = (self._current_price - self._sl_price) / total_range
            pos_ratio = max(0, min(1, pos_ratio))

            marker_x = rect.left() + (rect.width() * pos_ratio)

            # Draw marker line
            painter.setPen(QColor("#ffffff"))
            painter.drawLine(int(marker_x), rect.top() + 2, int(marker_x), rect.bottom() - 2)

            # Draw marker triangle
            painter.setBrush(QColor("#ffffff"))
            triangle_size = 6
            points = [
                (int(marker_x), rect.top()),
                (int(marker_x - triangle_size), rect.top() - triangle_size),
                (int(marker_x + triangle_size), rect.top() - triangle_size),
            ]
            from PyQt6.QtGui import QPolygon
            from PyQt6.QtCore import QPoint
            polygon = QPolygon([QPoint(x, y) for x, y in points])
            painter.drawPolygon(polygon)

        # Draw text
        painter.setPen(QColor("#ffffff"))
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, self.format())

        painter.end()

    def reset_bar(self):
        """Reset bar to default state."""
        self._sl_price = 0.0
        self._tp_price = 0.0
        self._current_price = 0.0
        self._entry_price = 0.0
        self.setValue(50)
        self.setFormat("No active position")
        self.update()

class BotUISignalsMixin:
    """BotUISignalsMixin extracted from BotUIPanelsMixin."""
    def _create_signals_tab(self) -> QWidget:
        """Create signals & trade management tab."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        # Top row: Bitunix HEDGE (with integrated Status) + Current Position
        top_row_layout = QHBoxLayout()
        top_row_layout.setSpacing(8)

        # Bitunix HEDGE Execution Widget (takes remaining space)
        # Now includes all 4 GroupBoxes: Connection & Risk, Entry, TP/SL, Status
        bitunix_widget = self._build_bitunix_hedge_widget()
        top_row_layout.addWidget(bitunix_widget, stretch=1)

        # Current Position (fixed 420px width)
        position_widget = self._build_current_position_widget()
        position_widget.setMaximumWidth(420)
        position_widget.setMinimumWidth(420)
        top_row_layout.addWidget(position_widget, stretch=0)

        layout.addLayout(top_row_layout)

        # 20px Abstand vor Recent Signals
        layout.addSpacing(10) # Reduced spacing

        # Issue #68: Use QSplitter for Signals Table and Bot Log
        # "Log Trading Bot ... nach unten hin minimieren kann"
        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.addWidget(self._build_signals_widget())
        splitter.addWidget(self._build_bot_log_widget())
        splitter.setStretchFactor(0, 7) # Signals table gets more space
        splitter.setStretchFactor(1, 3) # Log gets less
        
        layout.addWidget(splitter)
        
        return widget

    def _build_bitunix_hedge_widget(self) -> QWidget:
        """Build Bitunix HEDGE Execution Widget.

        Uses the new BitunixHedgeExecutionWidget with 3-column layout.
        Based on: UI_Spezifikation_TradingBot_Signals_BitunixExecution.md
        """
        from src.ui.widgets.bitunix_hedge_execution_widget import BitunixHedgeExecutionWidget

        try:
            self.bitunix_hedge_widget = BitunixHedgeExecutionWidget(parent=self)

            # Wire up signals
            self.bitunix_hedge_widget.order_placed.connect(self._on_bitunix_order_placed)
            self.bitunix_hedge_widget.position_opened.connect(self._on_bitunix_position_opened)
            self.bitunix_hedge_widget.trade_closed.connect(self._on_bitunix_trade_closed)

            return self.bitunix_hedge_widget

        except Exception as e:
            logger.error(f"Failed to create Bitunix HEDGE widget: {e}")
            # Return placeholder on error
            error_widget = QLabel(f"Bitunix HEDGE: Initialization failed - {e}")
            error_widget.setStyleSheet("color: #ff5555; padding: 8px;")
            return error_widget

    def _on_bitunix_order_placed(self, order_id: str):
        """Handle Bitunix order placed event."""
        logger.info(f"Bitunix order placed: {order_id}")

    def _on_bitunix_position_opened(self, position_id: str):
        """Handle Bitunix position opened event."""
        logger.info(f"Bitunix position opened: {position_id}")

    def _on_bitunix_trade_closed(self):
        """Handle Bitunix trade closed event."""
        logger.info("Bitunix trade closed")

    def _build_status_widget_fallback(self) -> QWidget:
        """Build fallback Status GroupBox if Bitunix widget failed."""
        status_group = QGroupBox("Status")
        layout = QFormLayout()
        layout.setContentsMargins(8, 12, 8, 8)
        layout.setVerticalSpacing(8)

        self.state_label = QLabel("—")
        layout.addRow("State:", self.state_label)

        self.order_id_label = QLabel("—")
        layout.addRow("Order ID:", self.order_id_label)

        self.position_id_label = QLabel("—")
        layout.addRow("Position ID:", self.position_id_label)

        self.adaptive_label = QLabel("—")
        layout.addRow("Adaptive:", self.adaptive_label)

        kill_btn = QPushButton("KILL SWITCH")
        kill_btn.setStyleSheet(
            "background-color: #ff0000; color: white; font-weight: bold; padding: 8px;"
        )
        layout.addRow(kill_btn)

        status_group.setLayout(layout)
        status_group.setMinimumWidth(160)
        status_group.setMaximumWidth(200)

        return status_group

    def _build_current_position_widget(self) -> QWidget:
        position_widget = QWidget()
        position_layout = QVBoxLayout(position_widget)
        position_layout.setContentsMargins(0, 0, 0, 0)
        position_layout.setSpacing(0)

        position_group = QGroupBox("Current Position")
        group_layout = QVBoxLayout()
        group_layout.setContentsMargins(6, 6, 6, 6)
        group_layout.setSpacing(4)

        # Issue #68: Add Live Trading controls here
        if hasattr(self, 'bitunix_hedge_widget'):
            mode_layout = QHBoxLayout()
            mode_layout.setContentsMargins(0, 0, 0, 4)
            mode_layout.setSpacing(12)
            
            # Use widgets from bitunix widget that were initialized but not added to layout
            cb = self.bitunix_hedge_widget.live_mode_cb
            lbl = self.bitunix_hedge_widget.mode_indicator
            
            # Adjust label style (smaller font as requested)
            lbl.setStyleSheet(
                "background-color: #26a69a; color: white; "
                "font-weight: bold; font-size: 11px; " 
                "padding: 4px 8px; border-radius: 4px;"
                "min-width: 100px;"
            )
            
            mode_layout.addWidget(cb)
            mode_layout.addWidget(lbl)
            mode_layout.addStretch()
            
            group_layout.addLayout(mode_layout)

        # Add SL/TP Progress Bar at the top
        self.sltp_progress_bar = SLTPProgressBar()
        self.sltp_progress_bar.reset_bar()
        group_layout.addWidget(self.sltp_progress_bar)

        # Add position details
        details_widget = QWidget()
        details_widget.setLayout(self._build_position_layout())
        group_layout.addWidget(details_widget)

        position_group.setLayout(group_layout)
        # Removed setMaximumHeight - allow full vertical space
        position_layout.addWidget(position_group)
        return position_widget

    def _build_position_layout(self) -> QHBoxLayout:
        position_h_layout = QHBoxLayout()
        position_h_layout.setContentsMargins(8, 8, 8, 8)
        position_h_layout.setSpacing(20)

        position_h_layout.addWidget(self._build_position_left_column())
        position_h_layout.addWidget(self._build_position_right_column())
        position_h_layout.addStretch()
        return position_h_layout

    def _build_position_left_column(self) -> QWidget:
        left_widget = QWidget()
        left_widget.setMinimumWidth(180)
        left_widget.setMaximumWidth(220)
        left_form = QFormLayout(left_widget)
        left_form.setVerticalSpacing(2)
        left_form.setContentsMargins(0, 0, 0, 0)

        self.position_side_label = QLabel("FLAT")
        self.position_side_label.setStyleSheet("font-weight: bold;")
        left_form.addRow("Side:", self.position_side_label)

        self.position_strategy_label = QLabel("-")
        left_form.addRow("Strategy:", self.position_strategy_label)

        self.position_entry_label = QLabel("-")
        left_form.addRow("Entry:", self.position_entry_label)

        self.position_size_label = QLabel("-")
        left_form.addRow("Size:", self.position_size_label)

        self.position_invested_label = QLabel("-")
        self.position_invested_label.setStyleSheet("font-weight: bold;")
        left_form.addRow("Invested:", self.position_invested_label)

        self.position_stop_label = QLabel("-")
        left_form.addRow("Stop:", self.position_stop_label)

        self.position_current_label = QLabel("-")
        self.position_current_label.setStyleSheet("font-weight: bold;")
        left_form.addRow("Current:", self.position_current_label)

        self.position_pnl_label = QLabel("-")
        left_form.addRow("P&L:", self.position_pnl_label)

        self.position_bars_held_label = QLabel("-")
        left_form.addRow("Bars Held:", self.position_bars_held_label)
        return left_widget

    def _build_position_right_column(self) -> QWidget:
        right_widget = QWidget()
        right_widget.setMinimumWidth(160)
        right_form = QFormLayout(right_widget)
        right_form.setVerticalSpacing(2)
        right_form.setContentsMargins(0, 0, 0, 0)

        self.position_score_label = QLabel("-")
        self.position_score_label.setStyleSheet("font-weight: bold; color: #26a69a;")
        right_form.addRow("Score:", self.position_score_label)

        self.position_tr_price_label = QLabel("-")
        right_form.addRow("TR Kurs:", self.position_tr_price_label)

        self.deriv_separator = QLabel("── Derivat ──")
        self.deriv_separator.setStyleSheet("color: #ff5722; font-weight: bold;")
        right_form.addRow(self.deriv_separator)

        self.deriv_wkn_label = QLabel("-")
        self.deriv_wkn_label.setStyleSheet("font-weight: bold;")
        right_form.addRow("WKN:", self.deriv_wkn_label)

        self.deriv_leverage_label = QLabel("-")
        right_form.addRow("Hebel:", self.deriv_leverage_label)

        self.deriv_spread_label = QLabel("-")
        right_form.addRow("Spread:", self.deriv_spread_label)

        self.deriv_ask_label = QLabel("-")
        right_form.addRow("Ask:", self.deriv_ask_label)

        self.deriv_pnl_label = QLabel("-")
        self.deriv_pnl_label.setStyleSheet("font-weight: bold;")
        right_form.addRow("D P&L:", self.deriv_pnl_label)
        return right_widget

    def _build_signals_widget(self) -> QWidget:
        signals_widget = QWidget()
        signals_layout = QVBoxLayout(signals_widget)
        signals_layout.setContentsMargins(0, 0, 0, 0)

        signals_group = QGroupBox("Recent Signals")
        signals_inner = QVBoxLayout()

        # === TOOLBAR mit Clear-Buttons ===
        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(0, 0, 0, 4)

        self.clear_selected_btn = QPushButton("🗑️ Zeile löschen")
        self.clear_selected_btn.setFixedHeight(24)
        self.clear_selected_btn.setStyleSheet("font-size: 10px; padding: 2px 8px;")
        self.clear_selected_btn.setToolTip("Ausgewählte Zeile löschen")
        self.clear_selected_btn.clicked.connect(self._on_clear_selected_signal)
        toolbar.addWidget(self.clear_selected_btn)

        self.clear_all_signals_btn = QPushButton("🧹 Alle löschen")
        self.clear_all_signals_btn.setFixedHeight(24)
        self.clear_all_signals_btn.setStyleSheet(
            "font-size: 10px; padding: 2px 8px; color: #ef5350;"
        )
        self.clear_all_signals_btn.setToolTip("Alle Signale aus der Tabelle löschen")
        self.clear_all_signals_btn.clicked.connect(self._on_clear_all_signals)
        toolbar.addWidget(self.clear_all_signals_btn)

        # Issue #11: Sofort Verkaufen Button
        self.sell_position_btn = QPushButton("💰 Sofort verkaufen")
        self.sell_position_btn.setFixedHeight(24)
        self.sell_position_btn.setStyleSheet(
            "font-size: 10px; padding: 2px 8px; background-color: #ff5722; color: white; font-weight: bold;"
        )
        self.sell_position_btn.setToolTip(
            "Aktuelle Position mit Limit-Order verkaufen (0.05% unter aktuellem Kurs)"
        )
        self.sell_position_btn.clicked.connect(self._on_sell_position_clicked)
        self.sell_position_btn.setEnabled(False)  # Initially disabled
        toolbar.addWidget(self.sell_position_btn)

        # Issue #18: Chart-Elemente zeichnen Button
        self.draw_chart_elements_btn = QPushButton("📊 Chart-Elemente")
        self.draw_chart_elements_btn.setFixedHeight(24)
        self.draw_chart_elements_btn.setStyleSheet(
            "font-size: 10px; padding: 2px 8px; background-color: #2196f3; color: white;"
        )
        self.draw_chart_elements_btn.setToolTip(
            "Zeichnet Entry, Stop-Loss und Trailing-Stop Linien für aktive Position im Chart.\n"
            "Linien können im Chart verschoben werden - Werte werden automatisch aktualisiert."
        )
        self.draw_chart_elements_btn.clicked.connect(self._on_draw_chart_elements_clicked)
        self.draw_chart_elements_btn.setEnabled(False)  # Initially disabled
        toolbar.addWidget(self.draw_chart_elements_btn)

        # Export to XLSX button (neben Chart-Elemente Button)
        self.export_signals_btn = QPushButton("📊 Export XLSX")
        self.export_signals_btn.setFixedHeight(24)
        self.export_signals_btn.setStyleSheet(
            "font-size: 10px; padding: 2px 8px; background-color: #2196f3; color: white;"
        )
        self.export_signals_btn.setToolTip("Export signal table to Excel file")
        self.export_signals_btn.clicked.connect(self._export_signals_to_xlsx)
        toolbar.addWidget(self.export_signals_btn)

        toolbar.addStretch()

        signals_inner.addLayout(toolbar)
        self._build_signals_table()
        signals_inner.addWidget(self.signals_table)

        signals_group.setLayout(signals_inner)
        signals_layout.addWidget(signals_group)
        return signals_widget

    # =========================================================================
    # Trading Bot Log (Issue #23)
    # =========================================================================

    def _append_bot_log(self, log_type: str, message: str, timestamp: str | None = None) -> None:
        """Append a log line to the Trading Bot Log UI."""
        if not hasattr(self, 'bot_log_text') or self.bot_log_text is None:
            return
        ts = timestamp or datetime.now().strftime("%H:%M:%S")
        entry = f"[{ts}] [{log_type.upper()}] {message}"
        self.bot_log_text.appendPlainText(entry)

    def _set_bot_run_status_label(self, running: bool) -> None:
        if not hasattr(self, 'bot_run_status_label'):
            return
        color = "#26a69a" if running else "#9e9e9e"
        state = "RUNNING" if running else "STOPPED"
        self.bot_run_status_label.setText(f"Status: {state}")
        self.bot_run_status_label.setStyleSheet(f"font-weight: bold; color: {color};")

    def _clear_bot_log(self) -> None:
        if hasattr(self, 'bot_log_text'):
            self.bot_log_text.clear()

    def _save_bot_log(self) -> None:
        if not hasattr(self, 'bot_log_text'):
            return
        content = self.bot_log_text.toPlainText()
        if not content.strip():
            QMessageBox.information(self, "Keine Logs", "Es sind keine Log-Einträge vorhanden.")
            return
        default_name = f"trading_bot_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Log speichern",
            default_name,
            "Text Files (*.txt);;Markdown (*.md);;All Files (*)",
        )
        if not file_path:
            return
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            QMessageBox.information(self, "Gespeichert", f"Log gespeichert: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Log konnte nicht gespeichert werden: {e}")

    def _build_bot_log_widget(self) -> QWidget:
        """Create Trading Bot Log group for Signals tab (Issue #23)."""
        group = QGroupBox("Log Trading Bot")
        layout = QVBoxLayout(group)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        # Status label
        status_row = QHBoxLayout()
        self.bot_run_status_label = QLabel("Status: STOPPED")
        self.bot_run_status_label.setStyleSheet("font-weight: bold; color: #9e9e9e;")
        status_row.addWidget(self.bot_run_status_label)
        status_row.addStretch()
        layout.addLayout(status_row)

        # Log text field
        self.bot_log_text = QPlainTextEdit()
        self.bot_log_text.setReadOnly(True)
        self.bot_log_text.setPlaceholderText("Bot-Aktivitäten, Status und Fehler werden hier protokolliert...")
        self.bot_log_text.setMinimumHeight(140)
        layout.addWidget(self.bot_log_text)

        # Buttons
        btn_row = QHBoxLayout()
        self.bot_log_save_btn = QPushButton("💾 Speichern")
        self.bot_log_save_btn.clicked.connect(self._save_bot_log)
        btn_row.addWidget(self.bot_log_save_btn)

        self.bot_log_clear_btn = QPushButton("🧹 Leeren")
        self.bot_log_clear_btn.clicked.connect(self._clear_bot_log)
        btn_row.addWidget(self.bot_log_clear_btn)

        btn_row.addStretch()
        layout.addLayout(btn_row)

        return group

    def _on_clear_selected_signal(self) -> None:
        """Löscht die ausgewählte Zeile aus der Signals-Tabelle."""
        selected_rows = self.signals_table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # Von unten nach oben löschen, um Index-Verschiebung zu vermeiden
        rows_to_delete = sorted([idx.row() for idx in selected_rows], reverse=True)

        # Map table rows (latest on top) back to _signal_history indices
        history_len = len(self._signal_history)
        indices_to_delete = {
            history_len - 1 - row for row in rows_to_delete
            if 0 <= history_len - 1 - row < history_len
        }

        # Remove from history and UI
        if indices_to_delete:
            self._signal_history = [
                sig for i, sig in enumerate(self._signal_history)
                if i not in indices_to_delete
            ]
            self._save_signal_history()

        for row in rows_to_delete:
            self.signals_table.removeRow(row)

        # Refresh to ensure P&L / selection stays consistent
        self._update_signals_table()

    def _on_clear_all_signals(self) -> None:
        """Löscht alle Zeilen aus der Signals-Tabelle."""
        if self.signals_table.rowCount() == 0:
            return

        reply = QMessageBox.question(
            self,
            "Alle Signale löschen",
            f"Alle {self.signals_table.rowCount()} Signale aus der Tabelle löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            self._signal_history.clear()
            self._save_signal_history()
            self.signals_table.setRowCount(0)
            self._update_signals_table()

    def _on_sell_position_clicked(self) -> None:
        """Issue #11: Sell current position with limit order 0.05% below current price."""
        # Find active position
        active_sig = None
        if hasattr(self, '_find_active_signal'):
            active_sig = self._find_active_signal()

        if not active_sig:
            QMessageBox.warning(
                self,
                "Keine offene Position",
                "Es gibt keine offene Position zum Verkaufen.",
            )
            return

        # Get current price
        current_price = active_sig.get("current_price", 0)
        if current_price <= 0:
            current_price = active_sig.get("price", 0)

        if current_price <= 0:
            QMessageBox.warning(
                self,
                "Kein Kurs verfügbar",
                "Aktueller Kurs ist nicht verfügbar.",
            )
            return

        # Calculate limit price (0.05% below current for long, above for short)
        side = active_sig.get("side", "long").lower()
        current_price_decimal = Decimal(str(current_price))
        if side == "long":
            limit_price = current_price_decimal * Decimal("0.9995")  # 0.05% below
            order_side = "SELL"
        else:
            limit_price = current_price_decimal * Decimal("1.0005")  # 0.05% above for short
            order_side = "BUY"

        quantity = active_sig.get("quantity", 0)
        symbol = active_sig.get("symbol", "")

        if not symbol:
            # Try to get symbol from chart widget
            if hasattr(self, 'chart_widget') and hasattr(self.chart_widget, 'current_symbol'):
                symbol = self.chart_widget.current_symbol

        if quantity <= 0:
            QMessageBox.warning(
                self,
                "Keine Positionsgröße",
                "Die Positionsgröße ist nicht verfügbar.",
            )
            return

        # Confirmation dialog
        reply = QMessageBox.question(
            self,
            "Position verkaufen",
            f"Position verkaufen?\n\n"
            f"Symbol: {symbol}\n"
            f"Side: {order_side}\n"
            f"Menge: {quantity:.6f}\n"
            f"Aktueller Kurs: {current_price:.4f}\n"
            f"Limit-Preis: {limit_price:.4f} (0.05% {'unter' if side == 'long' else 'über'} Kurs)\n",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        # Execute the sell
        asyncio.create_task(self._execute_sell_order(
            symbol=symbol,
            side=order_side,
            quantity=quantity,
            limit_price=limit_price,
            signal=active_sig,
        ))

    async def _execute_sell_order(
        self,
        symbol: str,
        side: str,
        quantity: float,
        limit_price: Decimal,
        signal: dict,
    ) -> None:
        """Execute the sell order via adapter.

        Issue #11: Places a limit order to close the position.

        Args:
            symbol: Trading symbol
            side: Order side (SELL for long, BUY for short)
            quantity: Position quantity
            limit_price: Limit price for the order
            signal: The signal dict containing position info
        """
        try:
            # Get adapter
            adapter = getattr(self, '_bitunix_adapter', None)
            if not adapter:
                # Try from bitunix widget
                bitunix_widget = getattr(self, '_bitunix_widget', None)
                if bitunix_widget:
                    adapter = getattr(bitunix_widget, 'adapter', None)

            if not adapter:
                logger.error("No adapter available for selling position")
                QMessageBox.critical(
                    self,
                    "Fehler",
                    "Kein Broker-Adapter verfügbar.",
                )
                return

            # Import order types
            from src.core.broker.broker_types import OrderRequest
            from src.database.models import OrderSide, OrderType as DBOrderType

            # Create order request
            order = OrderRequest(
                symbol=symbol,
                side=OrderSide.SELL if side == "SELL" else OrderSide.BUY,
                quantity=Decimal(str(quantity)),
                order_type=DBOrderType.LIMIT,
                limit_price=Decimal(str(limit_price)),
            )

            logger.info(
                f"Issue #11: Placing limit {side} order: {symbol} qty={quantity} @ {limit_price}"
            )

            # Place order
            response = await adapter.place_order(order)

            if response and response.broker_order_id:
                logger.info(f"Issue #11: Order placed successfully: {response.broker_order_id}")

                # Update signal status
                signal["status"] = "CLOSING"
                signal["exit_order_id"] = response.broker_order_id
                if hasattr(self, '_save_signal_history'):
                    self._save_signal_history()
                if hasattr(self, '_update_signals_table'):
                    self._update_signals_table()

                # Add to KI log
                if hasattr(self, '_add_ki_log_entry'):
                    self._add_ki_log_entry(
                        "SELL",
                        f"Limit-Order platziert: {symbol} {side} {quantity:.6f} @ {limit_price:.4f}"
                    )

                # Issue #13: Remove position lines from chart when selling
                if hasattr(self, 'chart_widget'):
                    try:
                        self.chart_widget.remove_stop_line("initial_stop")
                        self.chart_widget.remove_stop_line("trailing_stop")
                        self.chart_widget.remove_stop_line("entry_line")
                        if hasattr(self, '_add_ki_log_entry'):
                            self._add_ki_log_entry("CHART", "Linien entfernt (Verkauf eingeleitet)")
                    except Exception as e:
                        logger.error(f"Error removing chart lines: {e}")

                QMessageBox.information(
                    self,
                    "Order platziert",
                    f"Limit-Order wurde platziert.\n\n"
                    f"Order ID: {response.broker_order_id}\n"
                    f"Preis: {limit_price:.4f}",
                )
            else:
                logger.error("Issue #11: Order placement failed - no order ID returned")
                QMessageBox.warning(
                    self,
                    "Order fehlgeschlagen",
                    "Die Order konnte nicht platziert werden. Siehe Logs für Details.",
                )

        except Exception as e:
            logger.exception(f"Issue #11: Failed to place sell order: {e}")
            QMessageBox.critical(
                self,
                "Fehler",
                f"Fehler beim Platzieren der Order:\n{str(e)}",
            )

    def _update_sell_button_state(self) -> None:
        """Update the sell button enabled state based on open positions.

        Issue #11: Button is enabled only when there's an open position.
        Issue #18: Also updates draw_chart_elements_btn state.
        """
        has_open_position = False
        if hasattr(self, '_find_active_signal'):
            active_sig = self._find_active_signal()
            has_open_position = active_sig is not None

        if hasattr(self, 'sell_position_btn'):
            self.sell_position_btn.setEnabled(has_open_position)

        # Issue #18: Update chart elements button state
        if hasattr(self, 'draw_chart_elements_btn'):
            self.draw_chart_elements_btn.setEnabled(has_open_position)

    def _on_draw_chart_elements_clicked(self) -> None:
        """Issue #18: Draw chart elements (Entry, SL, TR) for active position.

        Zeichnet Entry-Linie, Stop-Loss und Trailing-Stop Linien im Chart.
        Labels zeigen die Werte in Prozent an.
        Linien können im Chart verschoben werden - Werte in Signals werden automatisch aktualisiert.
        """
        # Find active position
        active_sig = None
        if hasattr(self, '_find_active_signal'):
            active_sig = self._find_active_signal()

        if not active_sig:
            QMessageBox.warning(
                self,
                "Keine offene Position",
                "Es gibt keine offene Position für Chart-Elemente.",
            )
            return

        # Get position details
        entry_price = active_sig.get("price", 0)
        stop_price = active_sig.get("stop_price", 0)
        trailing_price = active_sig.get("trailing_stop_price", 0)
        side = active_sig.get("side", "long")
        initial_sl_pct = active_sig.get("initial_sl_pct", 0)
        trailing_pct = active_sig.get("trailing_stop_pct", 0)
        trailing_activation_pct = active_sig.get("trailing_activation_pct", 0)
        current_price = active_sig.get("current_price", entry_price)

        if entry_price <= 0:
            QMessageBox.warning(
                self,
                "Keine Entry-Daten",
                "Entry-Preis ist nicht verfügbar.",
            )
            return

        # Check for chart widget
        if not hasattr(self, 'chart_widget') or not self.chart_widget:
            QMessageBox.warning(
                self,
                "Kein Chart",
                "Chart-Widget ist nicht verfügbar.",
            )
            return

        try:
            # Calculate percentages if not available
            if initial_sl_pct <= 0 and stop_price > 0 and entry_price > 0:
                initial_sl_pct = abs((stop_price - entry_price) / entry_price) * 100

            if trailing_pct <= 0 and trailing_price > 0 and entry_price > 0:
                if side == "long":
                    trailing_pct = abs((entry_price - trailing_price) / entry_price) * 100
                else:
                    trailing_pct = abs((trailing_price - entry_price) / entry_price) * 100

            # Calculate TRA% (distance from current price)
            tra_pct = 0
            if trailing_price > 0 and current_price > 0:
                tra_pct = abs((current_price - trailing_price) / current_price) * 100

            # Draw Entry Line (blue)
            entry_label = f"Entry @ {entry_price:.2f}"
            self.chart_widget.add_stop_line(
                line_id="entry_line",
                price=entry_price,
                line_type="initial",
                color="#2196f3",  # Blue
                label=entry_label
            )

            # Draw Stop-Loss Line (red) with percentage
            if stop_price > 0:
                sl_label = f"SL @ {stop_price:.2f} ({initial_sl_pct:.2f}%)"
                self.chart_widget.add_stop_line(
                    line_id="initial_stop",
                    price=stop_price,
                    line_type="initial",
                    color="#ef5350",  # Red
                    label=sl_label
                )

            # Draw Trailing-Stop Line (orange) with percentage
            if trailing_price > 0:
                tr_is_active = active_sig.get("tr_active", False)
                if tr_is_active:
                    tr_label = f"TSL @ {trailing_price:.2f} ({trailing_pct:.2f}% / TRA: {tra_pct:.2f}%)"
                    color = "#ff9800"  # Orange when active
                else:
                    tr_label = f"TSL @ {trailing_price:.2f} ({trailing_pct:.2f}%) [inaktiv]"
                    color = "#888888"  # Gray when inactive

                self.chart_widget.add_stop_line(
                    line_id="trailing_stop",
                    price=trailing_price,
                    line_type="trailing",
                    color=color,
                    label=tr_label
                )

            # Log action
            if hasattr(self, '_add_ki_log_entry'):
                lines_drawn = ["Entry"]
                if stop_price > 0:
                    lines_drawn.append(f"SL ({initial_sl_pct:.2f}%)")
                if trailing_price > 0:
                    lines_drawn.append(f"TR ({trailing_pct:.2f}%)")
                self._add_ki_log_entry("CHART", f"Elemente gezeichnet: {', '.join(lines_drawn)}")

            logger.info(
                f"Issue #18: Chart elements drawn - Entry: {entry_price:.2f}, "
                f"SL: {stop_price:.2f} ({initial_sl_pct:.2f}%), "
                f"TR: {trailing_price:.2f} ({trailing_pct:.2f}%)"
            )

        except Exception as e:
            logger.exception(f"Issue #18: Failed to draw chart elements: {e}")
            QMessageBox.critical(
                self,
                "Fehler",
                f"Fehler beim Zeichnen der Chart-Elemente:\n{str(e)}",
            )

    def _build_signals_table(self) -> None:
        self.signals_table = QTableWidget()
        self.signals_table.setColumnCount(21)  # Added "Strategy" column
        self.signals_table.setHorizontalHeaderLabels(
            [
                "Time", "Type", "Strategy", "Side", "Entry", "Stop", "SL%", "TR%",
                "TRA%", "TR Lock", "Status", "Current", "P&L %", "P&L USDT",
                "Fees USDT",  # Issue #6: BitUnix fees in USDT
                "D P&L USDT", "D P&L %", "Heb", "WKN", "Score", "TR Stop",
            ]
        )
        # Hidden columns: D P&L USDT (15), D P&L % (16), Heb (17), WKN (18), Score (19)
        for col in [15, 16, 17, 18]:
            self.signals_table.setColumnHidden(col, True)
        self.signals_table.setColumnHidden(19, True)
        self.signals_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.signals_table.setAlternatingRowColors(True)
        self.signals_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.signals_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)

        self.signals_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.signals_table.customContextMenuRequested.connect(
            self._show_signals_context_menu
        )
        self.signals_table.cellChanged.connect(self._on_signals_table_cell_changed)
        self.signals_table.itemSelectionChanged.connect(
            self._on_signals_table_selection_changed
        )
        self._signals_table_updating = False

    def _update_leverage_column_visibility(self) -> None:
        """Issue #1: Show/hide leverage column based on leverage override state."""
        if not hasattr(self, 'signals_table'):
            return

        show_leverage = False
        if hasattr(self, 'get_leverage_override'):
            override_enabled, _ = self.get_leverage_override()
            show_leverage = override_enabled

        # Column 17 is "Heb" (Leverage) - mit Strategy-Spalte
        self.signals_table.setColumnHidden(17, not show_leverage)

    def _export_signals_to_xlsx(self) -> None:
        """Export signals table to XLSX file."""
        try:
            from datetime import datetime
            from PyQt6.QtWidgets import QFileDialog, QMessageBox

            # Ask user for save location
            default_filename = f"signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Export Signals to XLSX",
                default_filename,
                "Excel Files (*.xlsx)"
            )

            if not filepath:
                return  # User cancelled

            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                QMessageBox.warning(
                    self,
                    "Module nicht gefunden",
                    "openpyxl ist nicht installiert.\n\nBitte installieren Sie es mit:\npip install openpyxl"
                )
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Signals"

            # Get headers from table
            headers = []
            for col in range(self.signals_table.columnCount()):
                if not self.signals_table.isColumnHidden(col):
                    header_item = self.signals_table.horizontalHeaderItem(col)
                    headers.append(header_item.text() if header_item else f"Column {col}")

            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Write data
            visible_col_mapping = []
            for col in range(self.signals_table.columnCount()):
                if not self.signals_table.isColumnHidden(col):
                    visible_col_mapping.append(col)

            for row_idx in range(self.signals_table.rowCount()):
                for excel_col_idx, table_col_idx in enumerate(visible_col_mapping, start=1):
                    item = self.signals_table.item(row_idx, table_col_idx)
                    if item:
                        ws.cell(row=row_idx + 2, column=excel_col_idx, value=item.text())

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save workbook
            wb.save(filepath)

            QMessageBox.information(
                self,
                "Export erfolgreich",
                f"Signals wurden erfolgreich exportiert:\n{filepath}"
            )

        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(
                self,
                "Export Fehler",
                f"Fehler beim Exportieren der Signale:\n{str(e)}"
            )